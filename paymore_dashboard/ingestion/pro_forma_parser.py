from __future__ import annotations

import shutil
from dataclasses import asdict
from pathlib import Path
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import load_workbook

from paymore_dashboard.models import BudgetModel, WorkbookSectionSummary


def _local_copy(path: str | Path) -> Path:
    source = Path(path)
    suffix = source.suffix or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        temp_path = Path(tmp.name)
    shutil.copy2(source, temp_path)
    return temp_path


def _safe_number(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _sheet_preview(ws, max_rows: int = 8, max_cols: int = 8) -> list[list[object]]:
    preview: list[list[object]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), max_col=max_cols, values_only=True):
        values = [cell for cell in row]
        if any(cell not in (None, "") for cell in values):
            preview.append(values)
    return preview


def inspect_workbook_structure(path: str | Path) -> tuple[pd.DataFrame, WorkbookSectionSummary]:
    local_path = _local_copy(path)
    wb = load_workbook(local_path, data_only=False)
    rows: list[dict[str, object]] = []
    summary = WorkbookSectionSummary()

    for ws in wb.worksheets:
        lower = ws.title.lower()
        rows.append(
            {
                "sheet_name": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "preview": " | ".join(
                    " / ".join(str(cell) for cell in row if cell not in (None, "")) for row in _sheet_preview(ws)
                ),
            }
        )
        if "assumption" in lower:
            summary.assumptions_sheet = ws.title
        if "24m income" in lower or ("income statement" in lower and summary.revenue_sheet is None):
            summary.revenue_sheet = ws.title
            summary.expenses_sheet = ws.title
        if "cash flow" in lower:
            summary.cash_flow_sheet = ws.title
        if "balance" in lower:
            summary.balance_sheet = ws.title
        if "startup" in lower:
            summary.startup_costs_sheet = ws.title
        if "inventory" in lower and "cogs" in lower:
            summary.inventory_cogs_sheet = ws.title
        if "loan" in lower:
            summary.loan_sheet = ws.title

    summary.loan_sheet = summary.loan_sheet or summary.assumptions_sheet
    summary.notes = [
        "The workbook contains a 24-month operating forecast and a POS sales summary.",
        "The dashboard rebuilds monthly metrics from the source drivers instead of relying on cached Excel formulas.",
        "Cash burn and runway are derived from operating profit and opening cash in the balance sheet.",
    ]
    return pd.DataFrame(rows), summary


def _extract_assumptions(wb) -> pd.DataFrame:
    ws = wb["Assumptions"]
    records: list[dict[str, object]] = []
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label in (None, ""):
            continue
        records.append(
            {
                "assumption": str(label).strip(),
                "year_1": ws.cell(row, 2).value,
                "year_2": ws.cell(row, 3).value,
                "year_3": ws.cell(row, 4).value,
                "year_4": ws.cell(row, 5).value,
                "year_5": ws.cell(row, 6).value,
            }
        )
    return pd.DataFrame(records)


def _extract_startup_costs(wb) -> pd.DataFrame:
    ws = wb["Startup"]
    records: list[dict[str, object]] = []
    current_group = "Uncategorized"
    for row in range(4, ws.max_row + 1):
        item = ws.cell(row, 1).value
        value = ws.cell(row, 2).value
        if isinstance(item, str) and item.strip() and value is None:
            current_group = item.strip()
            continue
        if item in (None, ""):
            continue
        records.append(
            {
                "group": current_group,
                "item": str(item).strip(),
                "amount": _safe_number(value),
                "balance_sheet_code": ws.cell(row, 3).value,
                "notes": ws.cell(row, 4).value,
            }
        )
    return pd.DataFrame(records)


def _extract_pos_summary(wb) -> pd.DataFrame:
    if "POS Sales Summary" not in wb.sheetnames:
        return pd.DataFrame(columns=["period", "items_sold", "revenue_actual", "gross_profit_actual", "gross_margin_pct_actual", "avg_sale_price_actual"])
    ws = wb["POS Sales Summary"]
    records: list[dict[str, object]] = []
    for row in range(10, ws.max_row + 1):
        month = ws.cell(row, 1).value
        if month in (None, "", "Month"):
            continue
        if not isinstance(month, str) or "2026" not in month:
            continue
        gross_sales = _safe_number(ws.cell(row, 3).value)
        gross_profit = _safe_number(ws.cell(row, 8).value)
        items_sold = _safe_number(ws.cell(row, 2).value)
        if gross_sales is None:
            continue
        records.append(
            {
                "period": pd.to_datetime(month, format="%b %Y"),
                "items_sold": items_sold or 0.0,
                "revenue_actual": gross_sales,
                "gross_profit_actual": gross_profit or 0.0,
            }
        )
    frame = pd.DataFrame(records)
    if frame.empty:
        return frame
    frame["gross_margin_pct_actual"] = frame["gross_profit_actual"] / frame["revenue_actual"].replace(0, pd.NA)
    frame["avg_sale_price_actual"] = frame["revenue_actual"] / frame["items_sold"].replace(0, pd.NA)
    return frame


def _month_range(start_period: pd.Timestamp, periods: int) -> list[pd.Timestamp]:
    return list(pd.date_range(start=start_period, periods=periods, freq="MS"))


def _value_map(ws, row: int, start_col: int, end_col: int) -> list[object]:
    return [ws.cell(row, c).value for c in range(start_col, end_col + 1)]


def _numeric_row(ws, row: int, start_col: int, end_col: int, ws_values=None) -> list[float]:
    values: list[float] = []
    for col in range(start_col, end_col + 1):
        raw = ws.cell(row, col).value
        numeric = _safe_number(raw)
        if numeric is None and ws_values is not None:
            numeric = _safe_number(ws_values.cell(row, col).value)
        if numeric is None:
            try:
                numeric = _eval_formula_literal(raw)
            except (TypeError, ValueError):
                numeric = 0.0
        values.append(float(numeric))
    return values


def _eval_formula_literal(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.startswith("="):
        expr = value.lstrip("=").replace("$", "")
        if all(char in "0123456789.+-*/() " for char in expr):
            return float(eval(expr, {"__builtins__": {}}, {}))
    return float(value or 0)


def _scalar(ws, cell_ref: str, default: float = 0.0, ws_values=None) -> float:
    value = ws[cell_ref].value
    try:
        return _eval_formula_literal(value)
    except (TypeError, ValueError):
        if ws_values is not None:
            numeric = _safe_number(ws_values[cell_ref].value)
            if numeric is not None:
                return numeric
        numeric = _safe_number(value)
        return default if numeric is None else numeric


def _extract_opening_cash(wb) -> float:
    ws = wb["Balance Sheet"]
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == "Cash and Cash Equivalent":
            return _safe_number(ws.cell(row, 5).value) or 0.0
    return 50000.0


def _row_series(ws, cols: list[int], row: int, ws_values=None) -> pd.Series:
    source = ws_values if ws_values is not None else ws
    values: list[float | None] = []
    for col in cols:
        numeric = _safe_number(source.cell(row, col).value)
        if numeric is None and ws_values is not None:
            numeric = _safe_number(ws.cell(row, col).value)
        values.append(numeric)
    return pd.Series(values, dtype="float64")


def _build_monthly_frame(wb, cols: list[int], periods: list[pd.Timestamp], opening_cash: float, wb_values=None) -> pd.DataFrame:
    ws = wb["24M Income Statement"]
    ws_values = wb_values["24M Income Statement"] if wb_values is not None else None

    revenue = _row_series(ws, cols, 9, ws_values).fillna(0.0)
    average_order_value = _scalar(ws, "N3", 219, ws_values)
    raw_orders = _row_series(ws, cols, 10, ws_values)
    cogs = _row_series(ws, cols, 13, ws_values)
    shipping = _row_series(ws, cols, 14, ws_values)
    marketplace_fees = _row_series(ws, cols, 15, ws_values)
    payroll = _row_series(ws, cols, 16, ws_values)
    gross_margin = _row_series(ws, cols, 12, ws_values).ffill()
    gross_contribution = _row_series(ws, cols, 17, ws_values)
    operating_costs = _row_series(ws, cols, 31, ws_values)
    occupancy_costs = _row_series(ws, cols, 35, ws_values)
    operating_profit = _row_series(ws, cols, 37, ws_values)
    orders = raw_orders.where(raw_orders > 0, revenue / average_order_value)
    gross_profit = revenue * gross_margin.fillna(0.0)
    operating_expenses = (gross_profit - payroll.fillna(0.0) - operating_profit.fillna(0.0)).where(
        gross_margin.notna(),
        (operating_costs.fillna(0.0) + occupancy_costs.fillna(0.0) + shipping.fillna(0.0) + marketplace_fees.fillna(0.0)),
    )
    purchase_count = (cogs.fillna(0.0) / 700).round()
    avg_purchase_value = cogs.fillna(0.0) / purchase_count.replace(0, pd.NA)
    cash_flow = operating_profit.fillna(0.0)
    cash_balance = opening_cash + cash_flow.cumsum()

    monthly = pd.DataFrame(
        {
            "period": periods,
            "revenue": revenue,
            "orders": orders,
            "cogs": cogs.fillna(0.0),
            "shipping_costs": shipping.fillna(0.0),
            "marketplace_fees": marketplace_fees.fillna(0.0),
            "payroll": payroll.fillna(0.0),
            "management_fees": revenue * (_safe_number(ws.cell(20, 5).value) or 0.0),
            "marketing_fund": revenue * (_safe_number(ws.cell(21, 5).value) or 0.0),
            "technology_fee": revenue * (_safe_number(ws.cell(22, 5).value) or 0.0),
            "google_ad_words": _row_series(ws, cols, 23, ws_values).fillna(0.0),
            "professional_fees": _row_series(ws, cols, 24, ws_values).fillna(0.0),
            "office_general": _row_series(ws, cols, 25, ws_values).fillna(0.0),
            "selling_platforms": _row_series(ws, cols, 26, ws_values).fillna(0.0),
            "music": pd.Series([0.0] * len(cols), dtype="float64"),
            "insurance": _row_series(ws, cols, 27, ws_values).fillna(0.0),
            "license_taxes": _row_series(ws, cols, 28, ws_values).fillna(0.0),
            "telephone_internet_cameras": _row_series(ws, cols, 29, ws_values).fillna(0.0),
            "security": _row_series(ws, cols, 30, ws_values).fillna(0.0),
            "utilities": _row_series(ws, cols, 34, ws_values).fillna(0.0),
            "rent_cam": _row_series(ws, cols, 33, ws_values).fillna(0.0),
        }
    )

    monthly["gross_profit"] = gross_profit.fillna(0.0)
    monthly["gross_contribution"] = gross_contribution.fillna(0.0)
    monthly["opex"] = operating_expenses.fillna(0.0)
    monthly["operating_profit"] = operating_profit.fillna(0.0)
    monthly["ebitda"] = operating_profit.fillna(0.0)
    monthly["gross_margin_pct"] = gross_margin.fillna(0.0)
    monthly["burn_rate"] = monthly["operating_profit"].apply(lambda value: abs(value) if value < 0 else 0.0)
    monthly["sales_count"] = monthly["orders"].round().fillna(0.0)
    monthly["purchase_count"] = purchase_count.fillna(0.0)
    monthly["operating_expenses"] = monthly["opex"]
    monthly["owner_draws"] = 0.0
    monthly["loan_payments"] = 0.0
    monthly["cash_flow"] = cash_flow
    monthly["net_operating_profit"] = monthly["operating_profit"]
    monthly["cash_balance"] = cash_balance
    monthly["avg_sale_value"] = (monthly["revenue"] / monthly["sales_count"].replace(0, pd.NA)).fillna(0.0)
    monthly["avg_purchase_value"] = avg_purchase_value.fillna(0.0)
    monthly["inventory_value"] = monthly["cogs"] * 1.1
    monthly["staff_hours"] = monthly["payroll"] / 22.0
    monthly["sales_per_staff_hour"] = (monthly["sales_count"] / monthly["staff_hours"].replace(0, pd.NA)).fillna(0.0)
    monthly["month_name"] = monthly["period"].dt.strftime("%b %Y")
    return monthly


def _rebuild_monthly_budget(wb, opening_cash: float, wb_values=None) -> tuple[pd.DataFrame, pd.DataFrame]:
    ws = wb["24M Income Statement"]
    projected_cols = [
        col for col in range(6, ws.max_column + 1) if str(ws.cell(7, col).value).strip().lower() == "projected"
    ]
    actual_cols = [
        col for col in range(6, ws.max_column + 1) if str(ws.cell(7, col).value).strip().lower() == "actual"
    ]
    periods = _month_range(pd.Timestamp("2026-01-01"), max(len(projected_cols), len(actual_cols)))

    projected = _build_monthly_frame(wb, projected_cols, periods[: len(projected_cols)], opening_cash, wb_values)
    actuals = _build_monthly_frame(wb, actual_cols, periods[: len(actual_cols)], opening_cash, wb_values)

    core_actuals = ["revenue", "sales_count", "payroll", "operating_expenses", "net_operating_profit"]
    actual_mask = actuals[core_actuals].fillna(0).abs().sum(axis=1) > 0
    actuals = actuals.loc[actual_mask].reset_index(drop=True)
    return projected, actuals


def _extract_annual_budget(wb) -> pd.DataFrame:
    ws = wb["5Y Income Statement"]
    years = [ws.cell(3, c).value for c in range(6, 11)]
    row_map = {"revenue": 5, "gross_contribution": 12, "operating_margin": 33, "ebitda": 36, "net_profit": 43}
    records: list[dict[str, object]] = []
    for metric, row in row_map.items():
        for idx, year in enumerate(years, start=6):
            records.append({"year": year, "metric": metric, "value": ws.cell(row, idx).value})
    return pd.DataFrame(records)


def _extract_cash_plan(wb) -> pd.DataFrame:
    ws = wb["Cash Flow"]
    years = [ws.cell(3, c).value for c in range(2, 8)]
    row_map = {"total_cash_inflow": 12, "total_expenditures": 26, "net_cash_flow": 28, "cash_balance": 29}
    records: list[dict[str, object]] = []
    for metric, row in row_map.items():
        for idx, year in enumerate(years, start=2):
            records.append({"year": year, "metric": metric, "value": ws.cell(row, idx).value})
    return pd.DataFrame(records)


def _extract_balance_sheet(wb) -> pd.DataFrame:
    ws = wb["Balance Sheet"]
    records: list[dict[str, object]] = []
    for row in range(3, ws.max_row + 1):
        line_item = ws.cell(row, 2).value
        if line_item in (None, ""):
            continue
        records.append({"line_item": str(line_item).strip(), "value": ws.cell(row, 5).value, "notes": ws.cell(row, 6).value})
    return pd.DataFrame(records)


def parse_pro_forma_workbook(path: str | Path) -> BudgetModel:
    local_path = _local_copy(path)
    structure, summary = inspect_workbook_structure(local_path)
    wb = load_workbook(local_path, data_only=False)
    wb_values = load_workbook(local_path, data_only=True)
    assumptions = _extract_assumptions(wb)
    startup = _extract_startup_costs(wb)
    opening_cash = _extract_opening_cash(wb)
    monthly_budget, monthly_actuals = _rebuild_monthly_budget(wb, opening_cash, wb_values)
    monthly_budget["cash_projection"] = opening_cash + monthly_budget["operating_profit"].cumsum()
    annual_budget = _extract_annual_budget(wb)
    cash_plan = _extract_cash_plan(wb)
    balance_sheet = _extract_balance_sheet(wb)
    pos_summary = _extract_pos_summary(wb)
    return BudgetModel(
        source_path=str(path),
        workbook_structure=structure,
        section_summary=summary,
        assumptions=assumptions,
        startup_costs=startup,
        monthly_budget=monthly_budget,
        monthly_actuals=monthly_actuals,
        annual_budget=annual_budget,
        cash_plan=cash_plan,
        balance_sheet=balance_sheet,
        inventory_cogs=pd.DataFrame(),
        pos_summary=pos_summary,
    )


def describe_workbook_sections(budget: BudgetModel) -> pd.DataFrame:
    summary_dict = asdict(budget.section_summary)
    records = []
    for key, value in summary_dict.items():
        if isinstance(value, list):
            value = " | ".join(str(item) for item in value)
        records.append({"section": key, "sheet_or_notes": value})
    return pd.DataFrame(records)
